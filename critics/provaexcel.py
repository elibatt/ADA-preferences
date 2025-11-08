from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from io import BytesIO
import pandas as pd
from collections import defaultdict
'''
def create_excel_sheet_from_text(text, output_file, sheet_name):

    data = {}
    current_section = None

    # Parse the text
    for line in text.split('\n'):
        if line.strip().startswith('QUAD'):
            current_section = 'QUAD'
            data[current_section] = []
        elif line.strip().startswith('DFQUAD'):
            current_section = 'DFQUAD'
            data[current_section] = []
        elif line.strip().startswith('REB'):
            current_section = 'REB'
            data[current_section] = []
        elif line.strip().startswith('ENERGY'):
            current_section = 'ENERGY'
            data[current_section] = []
        elif line.strip():
            print(line)
            name, val = line.split(maxsplit=1)
            data[current_section].append((name, float(val)))

  
    # Convert data to DataFrame
    dfs = {}
    for section, values in data.items():
        df = pd.DataFrame(values, columns=['Name', section])
        df = df.set_index('Name')
        dfs[section] = df
    print(dfs)
    # Merge DataFrames
    
    result = pd.concat(dfs.values(), axis=1)
    result.to_latex
    #print(result)
    try:
        # Load existing workbook or create a new one if it doesn't exist
        try:
            workbook = load_workbook(output_file)
        except FileNotFoundError:
            workbook = Workbook()

        # Check if the sheet already exists
        if sheet_name in workbook.sheetnames:
            # If the sheet already exists, select it
            worksheet = workbook[sheet_name]
        else:
            # If the sheet doesn't exist, create a new one
            worksheet = workbook.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(dataframe_to_rows(result, index=True, header=True), 1):
          
            for c_idx,  value in enumerate(row, 1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        workbook.save(output_file)

    except Exception as e:
        print(f"Error: {e}")
'''    
def parse_text_to_table(text):
    lines = text.strip().split('\n')
    data = defaultdict(dict)
    current_category = None
    
    for line in lines:
        if not line.strip():
            continue
        
        parts = line.split()
        if len(parts) == 1:
            current_category = parts[0]  # This is a new column header
        else:
            key = parts[0]  # Row identifier
            value = float(parts[1])  # Corresponding value
            data[key][current_category] = value
    
    df = pd.DataFrame(data).T.fillna('')  # Convert to DataFrame and transpose
    df.index.name = 'Element'
    df.reset_index(inplace=True)
    return df

def save_to_excel(df, filename='output.xlsx', sheet_name='Sheet1'):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Data saved to {filename} in sheet {sheet_name}")
    
# Example usage:

text='''
QUAD
director 0.7777777777777778
writer 1.0
CillianMurphy 0.3333333333333333
RobertDowneyJr. 1.0
acting 1.0
politics 0.3333333333333333
morality 0.25
isolation 1.0
themes 0.7333333333333334
film 0.9230769230769231
DFQUAD
director 0.7777777777777778
writer 1.0
CillianMurphy 0.3333333333333333
RobertDowneyJr. 1.0
acting 1.0
politics 0.3333333333333333
morality 0.25
isolation 1.0
themes 0.5666666666666667
film 0.5
REB
director 0.7777777777777778
writer 1.0
CillianMurphy 0.3333333333333333
RobertDowneyJr. 1.0
acting 0.0
politics 0.3333333333333333
morality 0.25
isolation 1.0
themes 0.3600802044467316
film 0.9281730464296875
ENERGY
director 0.7777777777777778
writer 1.0
CillianMurphy 0.3333333333333333
RobertDowneyJr. 1.0
acting 0.64
politics 0.3333333333333333
morality 0.25
acting 0.64
politics 0.3333333333333333
morality 0.25
politics 0.3333333333333333
morality 0.25
morality 0.25
isolation 1.0
themes 0.5652830188679245
film 0.9608814365105898
'''

print(parse_text_to_table(text))
save_to_excel(parse_text_to_table(text), "C:/Users/elisa/Desktop/provapython.xlsx", 'def2')
#create_excel_sheet_from_text(text4, "C:/Users/elisa/Desktop/provapython.xlsx", 'def2')

def chiamametodi(text, directory, sheet):
    save_to_excel(parse_text_to_table(text), directory, sheet)